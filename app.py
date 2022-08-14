from fileinput import close
from controllers.main import *
from openpyxl import Workbook
import json

keys = []
wb = Workbook()
ws = wb.active


textSearch = input('Digite sua busca!: ')
print('') 

result_json(textSearch)

data = export_array()

#
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)
    close()


#
with open('data.json',encoding="utf8") as f:
    json_data = json.load(f)

#
for i in range(len(json_data)) :
	sub_obj = json_data[i]
	if i == 0 :
		keys = list(sub_obj.keys())
		for k in range(len(keys)) :
			# row or column index start from 1
			ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
	for j in range(len(keys)) :
		ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
wb.save(f"{textSearch}.xlsx")